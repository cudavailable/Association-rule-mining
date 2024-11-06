import os
import time
import openpyxl
from logger import Logger

def load_dataset(data_path):
	""" load data from a specific path """
	workbook = openpyxl.load_workbook(data_path)
	sheet = workbook['Sheet3']

	rows = sheet.max_row + 1
	cols = sheet.max_column + 1

	dataset = []
	for i in range(1, rows):
		temp = []
		for j in range(1, cols):
			cell = sheet.cell(i, j).value
			if cell:
				temp.append(cell)
		dataset.append(temp)

	return dataset

def create_C1(dataset):
	""" create an initial candidate set """
	C1 = set()
	for t in dataset:
		for item in t:
			item_set = frozenset([item]) # convert array into set
			C1.add(item_set)
	return C1

def is_apriori(Ck_item, Lksub1):
	for item in Ck_item:
		sub = Ck_item - frozenset([item])
		if sub not in Lksub1:
			# sub is not a k-1 frequent item
			return False

	return True

def create_Ck(Lksub1, k):
	""" create k-candidate set by Lk-1 """
	Ck = set()
	len_Lksub1 = len(Lksub1)
	list_Lksub1 = list(Lksub1)

	for i in range(len_Lksub1):
		for j in range(1, len_Lksub1):
			l1 = list(list_Lksub1[i])
			l2 = list(list_Lksub1[j])

			l1.sort()
			l2.sort()

			if l1[:k-2] == l2[:k-2]:
				new_l = list_Lksub1[i] | list_Lksub1[j]
				if is_apriori(new_l, Lksub1):
					Ck.add(new_l)

	return Ck

def generate_Lk_by_Ck(dataset, Ck, min_support, support_data):
	""" generate frequent item sets by candidate sets """
	Lk = set()
	item_count = {}
	for t in dataset:
		for item in Ck:
			if item.issubset(t):
				# count when item appears in t
				if item in item_count:
					item_count[item] += 1
				else:
					item_count[item] = 1

	t_num = float(len(dataset))
	for item in item_count.keys():
		# add items whose support >= min_support
		if item_count[item] / t_num >= min_support:
			Lk.add(item)
			support_data[item] = item_count[item] / t_num

	return Lk

def generate_L(dataset, k, min_support):
	""" generate all of frequent item sets and corresponding support """
	support_data = {} # save the support of different frequent items
	C1 = create_C1(dataset)
	L1 = generate_Lk_by_Ck(dataset, C1, min_support, support_data) # generate L1 by C1
	Lksub1 = L1.copy()
	L = []
	L.append(L1)

	for i in range(2, k+1):
		Ci = create_Ck(Lksub1, i)
		Li = generate_Lk_by_Ck(dataset, Ci, min_support, support_data)
		Lksub1 = Li.copy()
		L.append(Li)

	return L, support_data

def generate_big_rules(L, support_data, min_conf):
	""" generate strong-association-rules by frequent item sets """
	big_rules = []
	sub_set_list = []
	for i in range(len(L)):
		for freq_set in L[i]:
			for sub_set in sub_set_list:
				if sub_set.issubset(freq_set):
					conf = support_data[freq_set] / support_data[freq_set-sub_set]
					big_rule = (freq_set-sub_set, sub_set, conf, support_data[freq_set])
					if conf >= min_conf and (big_rule not in big_rules):
						big_rules.append(big_rule)
			sub_set_list.append(freq_set)

	return big_rules

class config:
	""" combinations of support and confidence """
	support_list = [0.05, 0.1, 0.2, 0.3]
	conf_list = [0.8, 0.85, 0.9, 0.95]

if __name__ == '__main__':
	# load your dataset and mine strong association rules
	data_path = r"./Database/dataset.xlsx"
	dataset = load_dataset(data_path)

	log_dir = r"./log"
	if log_dir is not None and not os.path.exists(log_dir):
		os.mkdir(log_dir)

	total = 1 # record how many experiments have been completed
	for min_support in config.support_list:
		for min_conf in config.conf_list:
			# create specific logger
			title = "shopping" + "_" + str(min_support) + "_" + str(min_conf) + ".txt"
			logger = Logger(os.path.join(log_dir, title))

			# association rules mining
			start_time = time.time()
			L, support_data = generate_L(dataset, k=3, min_support=min_support)
			big_rule_list = generate_big_rules(L, support_data, min_conf=min_conf)
			end_time = time.time()
			duration = end_time - start_time

			# save the mining result
			logger.write(f"Experiment #{total}  Min_Sup : {min_support}, Min_Conf : {min_conf}\n")
			logger.write("----------------------------------------------\n")
			logger.write("Hints:\n")
			logger.write("  1.The following data is organized in (I1, I2, min_conf, min_sup)\n")
			logger.write("  2.All decimals are rounded to 6 decimal places.\n")
			logger.write("----------------------------------------------\n")
			for i, big_rule in enumerate(big_rule_list):
				I1 = list(big_rule[0])
				I2 = list(big_rule[1])
				logger.write(f"{i+1}  {I1}  {I2}  {big_rule[2]:.6f}  {big_rule[3]:.6f}\n")

			logger.write("\n\nAssociation rules mining has completed.\n\n")
			logger.write(f"The time spending on Apriori (seconds): {duration:.6f}\n")
			total += 1
